#! /usr/local/bin/python3

# section for module import
import sys
import re	# to handle regular expression
import openpyxl

# section for class definition

START_ROW = 2
START_COL = 2
SHEET_NAME = 'case.plan'

def trans_xlsx(fn):
    nes = []
    cases = []
    
    wb = openpyxl.load_workbook(fn)
    sh = wb[SHEET_NAME]

    # find out all NEs
    end_col = START_COL+1
    while True:
        val = sh.cell(START_ROW, end_col).value
        if val:
            nes.append(val.split('.')[1])
            end_col += 1
        else: 
            break
    end_col -= 1 

    # find out all cases
    end_row = START_ROW + 1
    while True:
        val = sh.cell(end_row, START_COL).value
        if val:
            cases.append(val)
            end_row += 1
        else:
            break
    end_row -= 1

    print(f'data range row:col: [{START_ROW}:{START_COL}]-[{end_row}:{end_col}]')
    print(f'NE: {nes}')
    print(f'CASE: {cases}')

    '''
    for x in range(START_ROW, end_row):
        for y in range(START_COL,end_col):
           print(sh.cell(x,y).value,end='\t')
        print()
    '''

    newsheet = wb.create_sheet('ana.tmp')
    rolling_row , rolling_col = 2,2
    newsheet.cell(rolling_row,rolling_col).value = 'case'
    newsheet.cell(rolling_row,rolling_col+1).value = 'NE'
    newsheet.cell(rolling_row,rolling_col+2).value = 'plan'
    for row in range(START_ROW+1, end_row):
        for col in range(START_COL+1, end_col):
            if str(sh.cell(row,col).value).upper() != 'NA':
                rolling_row +=1
                rolling_col = 2
                newsheet.cell(rolling_row,rolling_col).value = cases[row-START_ROW-1]
                newsheet.cell(rolling_row,rolling_col+1).value = nes[col-START_COL-1]
                newsheet.cell(rolling_row,rolling_col+2).value = sh.cell(row,col).value
    wb.save(fn)
    print(f'{rolling_row-2} records analyzed!')

'''
openpyxl.load_workbook(fn)

sheet = workbook[sn]

sheet.ceel(row,col).value =

workbook.save()

wb = Workbook()
sheet = wb.active
sheet = wb.create_sheet('output')
ws.title = 'output'


for row in ws.iter_rows(min_row = START_ROW):
    for cell in row:
        print(cell)
'''


# here for unit test function
def Unit_Test(fn):
    trans_xlsx(fn)

# here for main code, usually call for Unit_Test
if __name__=='__main__':
    if len(sys.argv) < 2:
        print('Usage: transform.4.wf fn')
        exit()
    fn = sys.argv[1]
    print(f'Input file: {fn}')
    Unit_Test(fn)



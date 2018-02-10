#! /usr/local/bin/python3

# example file for excel operation
import random
import openpyxl

# if font operation required, pls include below import statement
from openpyxl.styles import Font

class WorkBook():
	''' WorkBook is to wrap the workbook. each instance is a workbook
	'''

	def __init__(self,filename=None):
		''' if filename not provided, a new workbook would be created
		'''
		self.status = False

		if not filename:
			self.filename = "temp.xls.file.%04d.xlsx" % random.randint(1,1000) 
			self.workbook = openpyxl.Workbook()
			self.workbook.create_sheet(title="Sheet1", index=0)
			self.workbook.save(self.filename)
		else:
			self.filename = filename
		self.load_excel()

	def load_excel(self):
		''' load excel sheets into the instance
		'''
		try:
			self.workbook = openpyxl.load_workbook(self.filename)
			self.status = True
			print('[%s] loaded. %d sheets inside: %s' % (self.filename , len(self.workbook.sheetnames),self.workbook.sheetnames))
		except FileNotFoundError:
			print('Load error: %s not found' % self.filename)

	def get_sheet(self, index=0):
		if not self.status:
			return
		self.sheet = self.workbook[self.workbook.sheetnames[index]]
		print('Sheet loaded, data range: [%d:%d]' % (self.sheet.max_row,self.sheet.max_column))
		print('Sammple value in [A1 / B3] : [%s / %s]' % ( self.sheet['A1'].value, self.sheet['B3'].value))

	def manipulate_and_save(self):
		''' sample code of data manipulation
		'''
		print("SET TITLE / DATA; CREATE NEW SHEET; SET DIMENSION AND FONT")
		self.sheet.title = 'SAMPLE TITLE'
		self.sheet['A1'] = 1999
		self.sheet['B3'] = 'aloha'
		for i in range(8,21):
			for j in range(1,11):
				self.sheet.cell(row = i, column = j).value = i*10+j
		self.workbook.create_sheet(title='new sheet' ,index = 1)
		self.sheet.column_dimensions['B'].width = 40
		self.sheet.row_dimensions[1].height = 30
		self.sheet['B3'].font = Font(sz=18,bold=True, italic=True)
		self.workbook.save(self.filename + '.out.xlsx')


# unit test code below
if __name__ == '__main__':

	print('Context - [openpyxl version]: %s' % openpyxl.__version__)

	# load smaple.xlsx
	wb = WorkBook('sample.xlsx')
	# create empty excel file
	wbtemp = WorkBook()

	# move to default sheet
	wb.get_sheet()
	wbtemp.get_sheet()

	wb.manipulate_and_save()
	wbtemp.manipulate_and_save()

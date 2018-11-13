#! /Users/steve.j.wang/sw.env/python.env/bin/python 
''' This module is used to manipulate data from excel
'''

# example file for excel operation
import random
import openpyxl

# if font operation required, pls include below import statement
# from openpyxl.styles import Font

TMP_FILE_EXTENTION = 'tmp.xlsx'
RANDOM_MAX = 9999
DEFAULT_SHEET_NAME = 'sw_1'

# excel data access index starts from 1

class WorkBook():
    ''' WorkBook is to wrap the workbook. each instance is a workbook
    '''

    def __init__(self,filename=None):
        ''' if filename not provided, a new workbook would be created
        '''
        self.load_status = False

        if not filename:        # if no filename specified, initiate a temp excel file
            self.filename = f"{random.randint(1,RANDOM_MAX):04d}.{TMP_FILE_EXTENTION}" 
            self.workbook = openpyxl.Workbook()
            self.workbook.create_sheet(title=DEFAULT_SHEET_NAME, index=0)
            self.workbook.save(self.filename)
        else:
            self.filename = filename
        self.load_excel()

    def load_excel(self):
        ''' load excel sheets into the instance
        '''
        try:
            self.workbook = openpyxl.load_workbook(self.filename,data_only=True)    #DATA ONLY to avoid getting formula data instead of real value
            self.load_status = True
            print(f'[{self.filename}] loaded. {len(self.workbook.sheetnames)} sheets inside: {self.workbook.sheetnames}' )
        except FileNotFoundError:
            print(f'Load error: {self.filename} not found')

    def get_sheet_with_name(self, sheetname=DEFAULT_SHEET_NAME):
        ''' this function is for sheet info retrival
        '''
        # print('Sammple value in [A1 / B3] : [%s / %s]' % ( self.sheet['A1'].value, self.sheet['B3'].value))
        if not self.load_status:
            print(f'workbook not loaded yet')
            return
        if sheetname in self.workbook.sheetnames:
            self.sheet = self.workbook[sheetname]
            print(f'Sheet loaded, data range: X- {self.sheet.max_column} Y- {self.sheet.max_row}' )
        else:
            print(f'Sheet name: {sheetname} not found')

    def read_sheet(self,sheetname=DEFAULT_SHEET_NAME,row_start=1, col_start=1):
        ''' this function is for sheet read
            row_start and col_start is in microsoft excel index which starts from 1
        '''
        lines = self.sheet.max_row
        cols = self.sheet.max_column
        for row in self.sheet.iter_rows(min_row=row_start, max_row=self.sheet.max_row, max_col=self.sheet.max_column):
            for cell in row[col_start-1:]:      # here convert microsoft excel index into python index
                print(f'{cell.value}',end='\t')
            print('')
        # sample code for cell: self.sheet.cell(row=i, column=j

    def get_data_from_sheet(self,sheetname=DEFAULT_SHEET_NAME,row_start=1, col_start=1):
        ''' this function is to get data from sheet
            row_start and col_start is in microsoft excel index which starts from 1
        '''
        sheet = self.workbook[sheetname]
        data_set = [ [cell.value for cell in row[col_start-1:]] for row in sheet.iter_rows(min_row=row_start, max_row=self.sheet.max_row, max_col=sheet.max_column)]
        # generate 2D list for the table and return it
        return data_set

    def manipulate_and_save(self):
        ''' sample code of data manipulation
        '''
        print("SET TITLE / DATA; CREATE NEW SHEET; SET DIMENSION AND FONT")
        self.sheet.title = 'SAMPLE TITLE'
        self.sheet['A1'] = 1999
        self.sheet['B3'] = 'aloha'
        for i in range(8,21):
            for j in range(1,11):
                self.sheet.cell(row = i, column = j).value = i*10+j
        self.workbook.create_sheet(title='new sheet' ,index = 1)
        self.sheet.column_dimensions['B'].width = 40
        self.sheet.row_dimensions[1].height = 30
        self.sheet['B3'].font = Font(sz=18,bold=True, italic=True)
        self.workbook.save(self.filename + '.out.xlsx')


# unit test code below
if __name__ == '__main__':

    print('Context - [openpyxl version]: %s' % openpyxl.__version__)

    # load smaple.xlsx
    wb = WorkBook('waterfall.mii.sa.0813.0821.xlsx')
    # create empty excel file
    # wbtemp = WorkBook()

    # move to default sheet
    wb.get_sheet_with_name('case.list')
    print(f'{wb.get_data_from_sheet("case.list",6,2)}')
    print(f'{wb.get_data_from_sheet("lab.core",6,2)}')
    # wbtemp.get_sheet_with_index()

    # wb.manipulate_and_save()
    # wbtemp.manipulate_and_save()

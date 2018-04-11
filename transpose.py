#! /usr/local/bin/python3

# section for module import
import sys
import re	# to handle regular expression
import openpyxl

# section for class definition

START_ROW = 2
START_COL = 2

# SHEET_NAME = 'case.plan'

def transpose(para_dict):
    ''' transpose excel table
    -Vn n cols taken as 1 item
    '''
    titles = []
    items = []
    
    print(f'Load file: {para_dict["FN"]}')
    wb = openpyxl.load_workbook(para_dict['FN'])
    sh = wb.active
    # find out all titles 
    end_col = START_COL
    while True:
        val = sh.cell(START_ROW, end_col).value
        if val:
            titles.append(val.split('.')[-1])
            end_col += 1
        else: 
            break
    item_title = list(titles[0:para_dict['V']])
    titles = titles[para_dict['V']:]

    # find out all items
    end_row = START_ROW + 1
    while True:
        vals = []
        for i in range(para_dict['V']):
            val = sh.cell(end_row, START_COL+i).value
            if val:
                vals.append(val)
            else:
                break
        if len(vals) == para_dict['V']:
            items.append(vals)
            end_row += 1
        else:
            break
    end_row -= 1

    print(f'titles: {len(titles)}')
    print(f'titles for item: {len(item_title)}')
    print(f'items: {len(items)}')

    newsheet = wb.create_sheet('transposed')
    rolling_row , rolling_col = 2,2
    for item in item_title:
        newsheet.cell(rolling_row,rolling_col).value = item 
        rolling_col += 1

    rolling_col_title = rolling_col
    rolling_col_value = rolling_col + 1
    newsheet.cell(rolling_row,rolling_col_title).value = 'title'
    newsheet.cell(rolling_row,rolling_col_value).value = 'value'
    value_start_row = START_ROW + 1 
    value_start_col = START_COL + para_dict['V']

    # write into the new tab with transposed data, here, title first
    for inx_item in range(len(items)):
        for inx_title in range(len(titles)):
            val = sh.cell(inx_item + value_start_row,inx_title + value_start_col).value
            if str(val).upper() != 'NA':
                rolling_row +=1
                rolling_col = 2
                for ite in items[inx_item]:
                    newsheet.cell(rolling_row,rolling_col).value = ite
                    rolling_col += 1
                newsheet.cell(rolling_row,rolling_col).value = titles[inx_title]
                newsheet.cell(rolling_row,rolling_col+1).value = val

    wb.save(para_dict['FN'])
    print(f'{len(titles)*len(items)} cells transposed!')
    print(f'{rolling_row - START_ROW} records left')

'''
wb = Workbook()
sheet = wb.active
'''

def parse_para(paras):
    '''parse parameters
    '''
    para_dict = {}
    for para in paras:
        if para[0] == '-':
            para_dict[para[1].upper()] = int(para[2:])
        else:
            para_dict['FN'] = para
    if 'V' not in para_dict:
        para_dict['V'] = 1
    print(f'Request: {para_dict}')
    return para_dict

# here for unit test function
def Unit_Test(parsed_para):
    pass

# here for main code, usually call for Unit_Test
if __name__=='__main__':
    if len(sys.argv) < 2:
        print('Usage: transpose [-Vn] fn')
        exit()
    parsed_para = parse_para(sys.argv[1:])
    transpose(parsed_para)


